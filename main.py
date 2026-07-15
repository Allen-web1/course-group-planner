from io import BytesIO
import math
import re
from typing import Any

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from ortools.sat.python import cp_model


app = FastAPI(title="고교학점제 교과편성 최적화 API", version="0.3.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)

REQUIRED_SHEETS = ["과목개설현황", "선택과목현황", "수강신청내역"]


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm(value: Any) -> str:
    return re.sub(r"\s+", "", text(value)).lower()


def number(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def filled(value: Any) -> bool:
    return value not in (None, "", 0, "0")


def header_index(headers: list[Any], candidates: list[str]) -> int:
    wanted = {norm(item) for item in candidates}
    return next((i for i, value in enumerate(headers) if norm(value) in wanted), -1)


def sheet_rows(workbook, name: str) -> list[list[Any]]:
    if name not in workbook.sheetnames:
        return []
    return [list(row) for row in workbook[name].iter_rows(values_only=True)]


def analyze_workbook(content: bytes, filename: str, include_private: bool = False) -> dict[str, Any]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(400, f"엑셀 파일을 읽을 수 없습니다: {exc}") from exc

    issues: list[dict[str, Any]] = []
    for name in REQUIRED_SHEETS:
        if name not in workbook.sheetnames:
            issues.append({
                "level": "error",
                "code": "MISSING_SHEET",
                "message": f"필수 시트 '{name}'이 없습니다.",
                "target": name,
            })

    opening = sheet_rows(workbook, "과목개설현황")
    headers = opening[0] if opening else []
    type_i = header_index(headers, ["선택구분"])
    term_i = header_index(headers, ["운영학기", "학기"])
    choose_i = header_index(headers, ["택O", "택수", "선택수"])
    credits_i = header_index(headers, ["학점"])
    min_i = header_index(headers, ["최소정원", "최소 정원"])
    max_i = header_index(headers, ["최대정원", "최대 정원"])
    course_columns = [
        i for i, value in enumerate(headers)
        if re.fullmatch(r"과목\d+", norm(value))
    ]

    courses: list[dict[str, Any]] = []
    group_number = 0
    for row in opening[1:]:
        if type_i < 0 or norm(row[type_i] if type_i < len(row) else None) != "선택":
            continue
        group_number += 1
        term = text(row[term_i] if term_i >= 0 and term_i < len(row) else None) or "학기 미지정"
        choose_count = max(1, number(row[choose_i] if choose_i >= 0 and choose_i < len(row) else None, 1))
        credits = max(0, number(row[credits_i] if credits_i >= 0 and credits_i < len(row) else None))
        minimum = max(1, number(row[min_i] if min_i >= 0 and min_i < len(row) else None, 10))
        maximum = max(minimum, number(row[max_i] if max_i >= 0 and max_i < len(row) else None, 36))
        for column in course_columns:
            base_name = text(row[column] if column < len(row) else None)
            if base_name:
                courses.append({
                    "name": base_name,
                    "base_name": base_name,
                    "term": term,
                    "group_id": str(group_number),
                    "group_name": f"선택{group_number}",
                    "choose_count": choose_count,
                    "credits": credits,
                    "minimum": minimum,
                    "maximum": maximum,
                    "requested_sections": 0,
                })

    status_rows = sheet_rows(workbook, "선택과목현황")
    status_headers = status_rows[0] if status_rows else []
    status_term_i = header_index(status_headers, ["운영학기", "학기"])
    status_start = max(0, header_index(status_headers, ["과목1"]))
    for row_index in range(1, len(status_rows), 2):
        course_row = status_rows[row_index]
        number_row = status_rows[row_index + 1] if row_index + 1 < len(status_rows) else []
        row_term = text(course_row[status_term_i] if status_term_i >= 0 and status_term_i < len(course_row) else None)
        for column in range(status_start, len(course_row)):
            base_name = text(course_row[column])
            if not base_name:
                continue
            definition = next((
                course for course in courses
                if course["base_name"] == base_name
                and (not row_term or course["term"] == row_term)
            ), None)
            if definition:
                definition["requested_sections"] = max(
                    0, number(number_row[column] if column < len(number_row) else None)
                )

    duplicate_names: dict[str, int] = {}
    for course in courses:
        duplicate_names[course["base_name"]] = duplicate_names.get(course["base_name"], 0) + 1
    for course in courses:
        if duplicate_names[course["base_name"]] > 1:
            course["name"] = f'{course["term"]} · {course["base_name"]}'

    applications = sheet_rows(workbook, "수강신청내역")
    application_headers = applications[0] if applications else []
    id_i = header_index(application_headers, ["학번", "학생ID", "연번"])
    class_i = header_index(application_headers, ["반", "구반"])
    number_i = header_index(application_headers, ["번호", "구번호"])
    name_i = header_index(application_headers, ["이름", "성명"])
    gender_i = header_index(application_headers, ["성별"])
    score_i = header_index(application_headers, ["성적", "평균", "석차등급"])
    note_i = header_index(application_headers, ["비고"])
    known = [i for i in [id_i, class_i, number_i, name_i] if i >= 0]
    choice_start = note_i + 1 if note_i >= 0 else (max(known) + 1 if known else 0)

    students: list[dict[str, Any]] = []
    for row_number, row in enumerate(applications[1:], start=2):
        student_id = text(row[id_i] if id_i >= 0 and id_i < len(row) else None)
        student_name = text(row[name_i] if name_i >= 0 and name_i < len(row) else None)
        if not student_id and not student_name:
            continue
        choices = [
            course["name"] for offset, course in enumerate(courses)
            if choice_start + offset < len(row) and filled(row[choice_start + offset])
        ]
        students.append({
            "student_id": student_id or f"AUTO-{row_number}",
            "name": student_name or "이름 미입력",
            "old_class": text(row[class_i] if class_i >= 0 and class_i < len(row) else None),
            "old_number": text(row[number_i] if number_i >= 0 and number_i < len(row) else None),
            "gender": text(row[gender_i] if gender_i >= 0 and gender_i < len(row) else None),
            "score": float(row[score_i]) if score_i >= 0 and score_i < len(row) and isinstance(row[score_i], (int, float)) else None,
            "choices": choices,
        })

    id_counts: dict[str, int] = {}
    for student in students:
        id_counts[student["student_id"]] = id_counts.get(student["student_id"], 0) + 1
    for student_id, count in id_counts.items():
        if count > 1:
            issues.append({
                "level": "error",
                "code": "DUPLICATE_ID",
                "message": f"학번 {student_id}가 {count}번 입력됐습니다.",
                "target": student_id,
            })

    groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for course in courses:
        groups.setdefault((course["term"], course["group_id"]), []).append(course)
    for student in students:
        for group_courses in groups.values():
            selected = sum(course["name"] in student["choices"] for course in group_courses)
            required = group_courses[0]["choose_count"]
            if selected != required:
                issues.append({
                    "level": "error",
                    "code": "CHOICE_COUNT",
                    "message": (
                        f'{student["name"]} 학생이 {group_courses[0]["group_name"]}에서 '
                        f"{selected}과목을 선택했습니다. {required}과목이 필요합니다."
                    ),
                    "target": student["student_id"],
                })

    if min_i < 0 or max_i < 0:
        issues.append({
            "level": "warning",
            "code": "DEFAULT_CAPACITY",
            "message": "최소·최대 정원 열이 없어 기본값 10명/36명을 적용했습니다.",
            "target": "과목개설현황",
        })

    original_classes = {student["old_class"] for student in students if student["old_class"]}
    grade_match = re.search(r"([1-3])\s*학년", filename)
    result = {
        "file_name": filename,
        "grade": f"{grade_match.group(1)}학년" if grade_match else "학년 미확인",
        "sheet_names": workbook.sheetnames,
        "student_count": len(students),
        "course_count": len(courses),
        "choice_group_count": len(groups),
        "original_class_count": len(original_classes),
        "courses": courses,
        "issues": issues,
        "error_count": sum(issue["level"] == "error" for issue in issues),
        "warning_count": sum(issue["level"] == "warning" for issue in issues),
    }
    if include_private:
        result["students"] = students
        result["groups"] = [
            {
                "term": key[0],
                "group_id": key[1],
                "courses": [course["name"] for course in value],
                "choose_count": value[0]["choose_count"],
                "credits": value[0]["credits"],
            }
            for key, value in groups.items()
        ]
    return result


def balanced_section_sizes(total: int, sections: int) -> list[int]:
    if sections <= 0:
        return []
    base, extra = divmod(total, sections)
    return [base + (1 if i < extra else 0) for i in range(sections)]


def optimize_core(data: dict[str, Any], time_limit: int = 45) -> dict[str, Any]:
    """학생 원반과 교과 분반을 하나의 원천 데이터로 생성한다.

    이 단계에서는 새 원반의 정원을 동일하게 맞추면서 선택 패턴이 같은 학생을
    가능한 한 가까운 원반에 모은다. 이후 이동수업 시간 묶음 최적화가 이 결과를
    입력으로 사용한다.
    """
    if data["error_count"]:
        return {
            "status": "INPUT_ERROR",
            "message": "입력 검증 오류를 먼저 수정해야 합니다.",
            "issues": data["issues"],
        }

    students = data["students"]
    courses = data["courses"]
    student_count = len(students)
    class_count = data["original_class_count"] or max(1, round(student_count / 30))
    room_limit = class_count + 2
    low_size = student_count // class_count
    high_size = math.ceil(student_count / class_count)

    enrollment: dict[str, list[int]] = {
        course["name"]: [i for i, student in enumerate(students) if course["name"] in student["choices"]]
        for course in courses
    }
    issues = list(data["issues"])
    section_counts: dict[str, int] = {}
    section_sizes: dict[str, list[int]] = {}
    for course in courses:
        count = len(enrollment[course["name"]])
        requested = course["requested_sections"]
        sections = requested or max(1, math.ceil(count / course["maximum"]))
        section_counts[course["name"]] = sections
        sizes = balanced_section_sizes(count, sections)
        section_sizes[course["name"]] = sizes
        if count == 0:
            issues.append({
                "level": "error", "code": "NO_ENROLLMENT",
                "message": f'{course["name"]} 선택 학생이 없습니다.', "target": course["name"],
            })
        if any(size > course["maximum"] for size in sizes):
            issues.append({
                "level": "error", "code": "MAX_CAPACITY",
                "message": f'{course["name"]}은 예정 {sections}개 분반으로 최대 정원 {course["maximum"]}명을 지킬 수 없습니다.',
                "target": course["name"],
            })
        if any(size < course["minimum"] for size in sizes):
            issues.append({
                "level": "error", "code": "MIN_CAPACITY",
                "message": f'{course["name"]}은 예정 {sections}개 분반을 그대로 만들면 최소 정원 {course["minimum"]}명 미만 분반이 생깁니다.',
                "target": course["name"],
            })

    hard_errors = [issue for issue in issues if issue["level"] == "error"]
    if hard_errors:
        return {
            "status": "INFEASIBLE_INPUT",
            "message": "예정 분반 수와 최소·최대 정원을 동시에 만족할 수 없습니다.",
            "issues": issues,
        }

    model = cp_model.CpModel()
    x = {(s, c): model.new_bool_var(f"x_{s}_{c}") for s in range(student_count) for c in range(class_count)}
    for s in range(student_count):
        model.add(sum(x[s, c] for c in range(class_count)) == 1)
    for c in range(class_count):
        size = sum(x[s, c] for s in range(student_count))
        model.add(size >= low_size)
        model.add(size <= high_size)

    # 성별 인원도 가능한 범위 안에서 고르게 유지한다.
    genders = sorted({student["gender"] for student in students if student["gender"]})
    for gender in genders:
        members = [s for s, student in enumerate(students) if student["gender"] == gender]
        lo, hi = len(members) // class_count, math.ceil(len(members) / class_count)
        for c in range(class_count):
            model.add(sum(x[s, c] for s in members) >= lo)
            model.add(sum(x[s, c] for s in members) <= hi)

    # 한 과목 선택자가 흩어진 원반 수를 줄이면 실제 이동수업 때 같은 교과의
    # 동시분반을 줄일 수 있다. 단, 학급 인원 균형은 위의 강제조건이 우선한다.
    active_vars = []
    for course_index, course in enumerate(courses):
        members = enrollment[course["name"]]
        if not members:
            continue
        for c in range(class_count):
            active = model.new_bool_var(f"active_{course_index}_{c}")
            for s in members:
                model.add(x[s, c] <= active)
            model.add(active <= sum(x[s, c] for s in members))
            active_vars.append(active)
    model.minimize(sum(active_vars))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = max(5, min(time_limit, 120))
    solver.parameters.num_search_workers = 8
    solver.parameters.random_seed = 20260715
    status = solver.solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return {
            "status": "NO_SOLUTION",
            "message": "학급 인원과 성별 균형 조건을 만족하는 새 원반을 찾지 못했습니다.",
            "issues": issues,
        }

    new_classes = [next(c for c in range(class_count) if solver.value(x[s, c])) + 1 for s in range(student_count)]
    class_members = {
        str(c + 1): [students[s]["student_id"] for s in range(student_count) if new_classes[s] == c + 1]
        for c in range(class_count)
    }

    # 같은 교과 안에서는 새 원반이 한쪽 분반에 지나치게 몰리지 않도록
    # 원반 번호를 순환시키며 정확히 예정 분반 수로 나눈다.
    student_sections: dict[tuple[int, str], int] = {}
    course_sections: list[dict[str, Any]] = []
    for course in courses:
        name = course["name"]
        members = sorted(enrollment[name], key=lambda s: (new_classes[s], students[s]["student_id"]))
        sizes = section_sizes[name]
        buckets = [[] for _ in sizes]
        order = sorted(range(len(sizes)), key=lambda i: (-sizes[i], i))
        cursor = 0
        for s in members:
            while len(buckets[order[cursor % len(order)]]) >= sizes[order[cursor % len(order)]]:
                cursor += 1
            section = order[cursor % len(order)]
            buckets[section].append(s)
            student_sections[s, name] = section + 1
            cursor += 1
        for section, bucket in enumerate(buckets, start=1):
            course_sections.append({
                "course": name,
                "base_name": course["base_name"],
                "term": course["term"],
                "group_id": course["group_id"],
                "credits": course["credits"],
                "section": section,
                "count": len(bucket),
                "minimum": course["minimum"],
                "maximum": course["maximum"],
                "students": [students[s]["student_id"] for s in bucket],
            })

    student_rows = []
    for s, student in enumerate(students):
        allocations = [
            {"course": course, "section": student_sections[s, course]}
            for course in student["choices"]
            if (s, course) in student_sections
        ]
        student_rows.append({
            "student_id": student["student_id"], "name": student["name"],
            "old_class": student["old_class"], "old_number": student["old_number"],
            "new_class": new_classes[s], "allocations": allocations,
        })

    assigned_pairs = sum(len(row["allocations"]) for row in student_rows)
    expected_pairs = sum(len(student["choices"]) for student in students)
    validation = {
        "student_count": student_count,
        "new_class_count": class_count,
        "room_limit": room_limit,
        "class_sizes": {key: len(value) for key, value in class_members.items()},
        "expected_student_course_pairs": expected_pairs,
        "assigned_student_course_pairs": assigned_pairs,
        "unassigned_count": expected_pairs - assigned_pairs,
        "duplicate_assignment_count": 0,
        "course_section_total": len(course_sections),
        "passed": assigned_pairs == expected_pairs,
    }
    return {
        "status": "PREPARED",
        "message": "새 원반과 예정 교과 분반 배정을 생성했습니다. 다음 단계에서 실제 시간 묶음을 최적화합니다.",
        "grade": data["grade"],
        "issues": issues,
        "validation": validation,
        "students": student_rows,
        "course_sections": course_sections,
    }


@app.get("/")
def root():
    return {
        "service": "course-group-planner-api",
        "version": "0.3.0",
        "status": "ready",
        "message": "교과편성 최적화 서버가 정상 작동 중입니다.",
    }


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/analyze")
async def analyze(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, ".xlsx 또는 .xlsm 파일만 업로드할 수 있습니다.")
    content = await file.read()
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(413, "파일 크기는 25MB 이하여야 합니다.")
    return analyze_workbook(content, file.filename)


@app.post("/optimize")
async def optimize(file: UploadFile = File(...), time_limit: int = 45):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, ".xlsx 또는 .xlsm 파일만 업로드할 수 있습니다.")
    content = await file.read()
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(413, "파일 크기는 25MB 이하여야 합니다.")
    data = analyze_workbook(content, file.filename, include_private=True)
    return optimize_core(data, time_limit=time_limit)
