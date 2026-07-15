from io import BytesIO
import re
from typing import Any

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook


app = FastAPI(title="고교학점제 교과편성 최적화 API", version="0.2.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)

REQUIRED_SHEETS = ["과목개설현황", "선택과목현황", "수강신청내역"]


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm(value: Any) -> str:
    return re.sub(r"\s+", "", text(value)).lower()


def number(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def filled(value: Any) -> bool:
    return value not in (None, "", 0, "0")


def header_index(headers: list[Any], candidates: list[str]) -> int:
    wanted = {norm(item) for item in candidates}
    return next((i for i, value in enumerate(headers) if norm(value) in wanted), -1)


def sheet_rows(workbook, name: str) -> list[list[Any]]:
    if name not in workbook.sheetnames:
        return []
    return [list(row) for row in workbook[name].iter_rows(values_only=True)]


def analyze_workbook(content: bytes, filename: str) -> dict[str, Any]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(400, f"엑셀 파일을 읽을 수 없습니다: {exc}") from exc

    issues: list[dict[str, Any]] = []
    for name in REQUIRED_SHEETS:
        if name not in workbook.sheetnames:
            issues.append({
                "level": "error",
                "code": "MISSING_SHEET",
                "message": f"필수 시트 '{name}'이 없습니다.",
                "target": name,
            })

    opening = sheet_rows(workbook, "과목개설현황")
    headers = opening[0] if opening else []
    type_i = header_index(headers, ["선택구분"])
    term_i = header_index(headers, ["운영학기", "학기"])
    choose_i = header_index(headers, ["택O", "택수", "선택수"])
    credits_i = header_index(headers, ["학점"])
    min_i = header_index(headers, ["최소정원", "최소 정원"])
    max_i = header_index(headers, ["최대정원", "최대 정원"])
    course_columns = [
        i for i, value in enumerate(headers)
        if re.fullmatch(r"과목\d+", norm(value))
    ]

    courses: list[dict[str, Any]] = []
    group_number = 0
    for row in opening[1:]:
        if type_i < 0 or norm(row[type_i] if type_i < len(row) else None) != "선택":
            continue
        group_number += 1
        term = text(row[term_i] if term_i >= 0 and term_i < len(row) else None) or "학기 미지정"
        choose_count = max(1, number(row[choose_i] if choose_i >= 0 and choose_i < len(row) else None, 1))
        credits = max(0, number(row[credits_i] if credits_i >= 0 and credits_i < len(row) else None))
        minimum = max(1, number(row[min_i] if min_i >= 0 and min_i < len(row) else None, 10))
        maximum = max(minimum, number(row[max_i] if max_i >= 0 and max_i < len(row) else None, 36))
        for column in course_columns:
            base_name = text(row[column] if column < len(row) else None)
            if base_name:
                courses.append({
                    "name": base_name,
                    "base_name": base_name,
                    "term": term,
                    "group_id": str(group_number),
                    "group_name": f"선택{group_number}",
                    "choose_count": choose_count,
                    "credits": credits,
                    "minimum": minimum,
                    "maximum": maximum,
                    "requested_sections": 0,
                })

    status_rows = sheet_rows(workbook, "선택과목현황")
    status_headers = status_rows[0] if status_rows else []
    status_term_i = header_index(status_headers, ["운영학기", "학기"])
    status_start = max(0, header_index(status_headers, ["과목1"]))
    for row_index in range(1, len(status_rows), 2):
        course_row = status_rows[row_index]
        number_row = status_rows[row_index + 1] if row_index + 1 < len(status_rows) else []
        row_term = text(course_row[status_term_i] if status_term_i >= 0 and status_term_i < len(course_row) else None)
        for column in range(status_start, len(course_row)):
            base_name = text(course_row[column])
            if not base_name:
                continue
            definition = next((
                course for course in courses
                if course["base_name"] == base_name
                and (not row_term or course["term"] == row_term)
            ), None)
            if definition:
                definition["requested_sections"] = max(
                    0, number(number_row[column] if column < len(number_row) else None)
                )

    duplicate_names: dict[str, int] = {}
    for course in courses:
        duplicate_names[course["base_name"]] = duplicate_names.get(course["base_name"], 0) + 1
    for course in courses:
        if duplicate_names[course["base_name"]] > 1:
            course["name"] = f'{course["term"]} · {course["base_name"]}'

    applications = sheet_rows(workbook, "수강신청내역")
    application_headers = applications[0] if applications else []
    id_i = header_index(application_headers, ["학번", "학생ID", "연번"])
    class_i = header_index(application_headers, ["반", "구반"])
    number_i = header_index(application_headers, ["번호", "구번호"])
    name_i = header_index(application_headers, ["이름", "성명"])
    note_i = header_index(application_headers, ["비고"])
    known = [i for i in [id_i, class_i, number_i, name_i] if i >= 0]
    choice_start = note_i + 1 if note_i >= 0 else (max(known) + 1 if known else 0)

    students: list[dict[str, Any]] = []
    for row_number, row in enumerate(applications[1:], start=2):
        student_id = text(row[id_i] if id_i >= 0 and id_i < len(row) else None)
        student_name = text(row[name_i] if name_i >= 0 and name_i < len(row) else None)
        if not student_id and not student_name:
            continue
        choices = [
            course["name"] for offset, course in enumerate(courses)
            if choice_start + offset < len(row) and filled(row[choice_start + offset])
        ]
        students.append({
            "student_id": student_id or f"AUTO-{row_number}",
            "name": student_name or "이름 미입력",
            "old_class": text(row[class_i] if class_i >= 0 and class_i < len(row) else None),
            "old_number": text(row[number_i] if number_i >= 0 and number_i < len(row) else None),
            "choices": choices,
        })

    id_counts: dict[str, int] = {}
    for student in students:
        id_counts[student["student_id"]] = id_counts.get(student["student_id"], 0) + 1
    for student_id, count in id_counts.items():
        if count > 1:
            issues.append({
                "level": "error",
                "code": "DUPLICATE_ID",
                "message": f"학번 {student_id}가 {count}번 입력됐습니다.",
                "target": student_id,
            })

    groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for course in courses:
        groups.setdefault((course["term"], course["group_id"]), []).append(course)
    for student in students:
        for group_courses in groups.values():
            selected = sum(course["name"] in student["choices"] for course in group_courses)
            required = group_courses[0]["choose_count"]
            if selected != required:
                issues.append({
                    "level": "error",
                    "code": "CHOICE_COUNT",
                    "message": (
                        f'{student["name"]} 학생이 {group_courses[0]["group_name"]}에서 '
                        f"{selected}과목을 선택했습니다. {required}과목이 필요합니다."
                    ),
                    "target": student["student_id"],
                })

    if min_i < 0 or max_i < 0:
        issues.append({
            "level": "warning",
            "code": "DEFAULT_CAPACITY",
            "message": "최소·최대 정원 열이 없어 기본값 10명/36명을 적용했습니다.",
            "target": "과목개설현황",
        })

    original_classes = {student["old_class"] for student in students if student["old_class"]}
    grade_match = re.search(r"([1-3])\s*학년", filename)
    return {
        "file_name": filename,
        "grade": f"{grade_match.group(1)}학년" if grade_match else "학년 미확인",
        "sheet_names": workbook.sheetnames,
        "student_count": len(students),
        "course_count": len(courses),
        "choice_group_count": len(groups),
        "original_class_count": len(original_classes),
        "courses": courses,
        "issues": issues,
        "error_count": sum(issue["level"] == "error" for issue in issues),
        "warning_count": sum(issue["level"] == "warning" for issue in issues),
    }


@app.get("/")
def root():
    return {
        "service": "course-group-planner-api",
        "version": "0.2.0",
        "status": "ready",
        "message": "교과편성 최적화 서버가 정상 작동 중입니다.",
    }


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/analyze")
async def analyze(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, ".xlsx 또는 .xlsm 파일만 업로드할 수 있습니다.")
    content = await file.read()
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(413, "파일 크기는 25MB 이하여야 합니다.")
    return analyze_workbook(content, file.filename)
